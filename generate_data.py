import random
from random import randrange
from openpyxl import Workbook
from datetime import timedelta, datetime

def random_date(start, end):
    """
    This function will return a random datetime between two datetime 
    objects.
    """
    delta = end - start
    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
    random_second = randrange(int_delta)
    return start + timedelta(seconds=random_second)

d1 = datetime.strptime('1/1/2023 1:00 PM', '%m/%d/%Y %I:%M %p')
d2 = datetime.strptime('4/25/2023 9:00 AM', '%m/%d/%Y %I:%M %p')

print(random_date(d1, d2))

# Define the columns and their data types
columns = {
    'Item_ID': int,
    'ItemName': str,
    'SectionID': int,
    'Date': str,
}

# Define the range of values for each column
ranges = {
    'Item_ID': (1, 13),
    'ItemName': ('Section Item 1', 'Section Item 2', 'Section Item 3', 'Section Item 4', 'Section Item 5', 'Section Item 6'),
    'SectionID':range(1, 1562)
}

# Generate fake data
data = []
for section_id in ranges['SectionID']:
    for i in range(1, 4):
        item_id = i
        discussion_name = random.choice(ranges['ItemName'])
        active = random.choice(ranges['SectionID'])
        date=random_date(d1, d2)
        data.append((item_id, discussion_name, section_id, date))

# Create a new workbook and sheet
wb = Workbook()
ws = wb.active

# Add the column headers
for i, col in enumerate(columns.keys()):
    ws.cell(row=1, column=i+1, value=col)

# Add the data to the sheet
for i, row_data in enumerate(data):
    for j, col in enumerate(columns.keys()):
        ws.cell(row=i+2, column=j+1, value=row_data[j])

# Save the workbook to an XLSX file
wb.save("dates.xlsx")